import streamlit as st
import pandas as pd
import numpy as np
from datetime import date
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side, Alignment
from openpyxl.utils import get_column_letter
import re
from pathlib import Path
from typing import Optional, Dict

st.set_page_config(page_title="LIQUIDACIONES (CSV rules)", page_icon="🏦", layout="wide")

# ---------- Constantes / utilidades ----------
MONEY_COLS_CANON = {
    "Ingreso alojamiento","Ingreso limpieza","Total ingresos","Comisión portal",
    "Honorarios Florit","Gasto limpieza","Amenities","Total Gastos","Pago al propietario",
    "Pago recibido","IVA del alquiler"
}
NIGHTS_COL = "Noches ocupadas"

def ensure_unique_columns(df: pd.DataFrame) -> pd.DataFrame:
    counts, new_cols = {}, []
    for c in df.columns:
        name = str(c)
        n = counts.get(name, 0)
        new_cols.append(name if n == 0 else f"{name}.{n}")
        counts[name] = n + 1
    out = df.copy()
    out.columns = new_cols
    return out

def _first_existing(df, candidates):
    norm_map = {str(c).strip().lower(): c for c in df.columns}
    for cand in candidates:
        k = str(cand).strip().lower()
        if k in norm_map:
            return norm_map[k]
    return None

def ensure_required(df, required, ctx=""):
    missing = [c for c in required if c not in df.columns]
    if missing:
        st.error(f"Faltan columnas requeridas: {missing} en {ctx}. Ajusta el archivo o usa el modo por letras.")
        st.stop()

def base_name(colname: str) -> str:
    return re.sub(r"\.\d+$", "", str(colname)).strip()

def is_money_col(colname: str) -> bool:
    return base_name(colname) in MONEY_COLS_CANON

def is_nights_col(colname: str) -> bool:
    return base_name(colname).lower() == NIGHTS_COL.lower()

def fmt_number_for_ui(colname: str, x):
    if is_nights_col(colname):
        try: return f"{int(round(float(x)))}"
        except Exception: return x
    if is_money_col(colname):
        try:
            s = f"{float(x):,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
            return f"{s} €"
        except Exception:
            return x
    try:
        return f"{float(x):,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
    except Exception:
        return x

def find_col(df: pd.DataFrame, base: str):
    for c in df.columns:
        if base_name(c).lower() == base.strip().lower():
            return c
    return None

def show_table_es_grouped(df: pd.DataFrame, title: str, group_col: str = "Alojamiento"):
    st.subheader(title)
    if group_col not in df.columns:
        view = df.copy()
        total = {c: (view[c].sum() if pd.api.types.is_numeric_dtype(view[c]) else "") for c in view.columns}
        view = pd.concat([view, pd.DataFrame([total], index=["TOTAL"])], axis=0)
        view_fmt = view.copy()
        for c in view_fmt.columns:
            if pd.api.types.is_numeric_dtype(view[c]) or is_money_col(c) or is_nights_col(c):
                view_fmt[c] = view_fmt[c].apply(lambda v: fmt_number_for_ui(c, v))
        def highlight_total(row):
            return ["font-weight: bold;" if row.name == "TOTAL" else "" for _ in row]
        st.dataframe(view_fmt.style.apply(highlight_total, axis=1), use_container_width=True)
        return
    for aloj, subdf in df.groupby(group_col):
        st.markdown(f"**{aloj}**")
        block = subdf.copy()
        total = {c: (block[c].sum() if pd.api.types.is_numeric_dtype(block[c]) else "") for c in block.columns}
        block = pd.concat([block, pd.DataFrame([total], index=["TOTAL"])], axis=0)
        block_fmt = block.copy()
        for c in block_fmt.columns:
            if pd.api.types.is_numeric_dtype(block[c]) or is_money_col(c) or is_nights_col(c):
                block_fmt[c] = block_fmt[c].apply(lambda v: fmt_number_for_ui(c, v))
        def highlight_total(row):
            return ["font-weight: bold;" if row.name == "TOTAL" else "" for _ in row]
        st.dataframe(block_fmt.style.apply(highlight_total, axis=1), use_container_width=True)
        st.divider()

# ---------- Normalización de columnas ----------
LETTER_MAP_DEFAULT = {
    "W": "Alojamiento",
    "D": "Fecha entrada",
    "F": "Fecha salida",
    "H": "Noches ocupadas",
    "I": "Ingreso alojamiento",
    "L": "Ingreso limpieza",
    "O": "Total ingresos",
    "AP": "Portal",
    "AR": "Comisión portal",
    "AL": "IVA del alquiler",
}

def letters_to_idx(letter):
    s = letter.upper(); n = 0
    for ch in s:
        if not ('A' <= ch <= 'Z'): return None
        n = n*26 + (ord(ch)-ord('A')+1)
    return n-1

def normalize_columns_by_letters(df, letter_map=LETTER_MAP_DEFAULT):
    out = df.copy(); cols = list(out.columns); rename = {}
    for L, std in letter_map.items():
        i = letters_to_idx(L)
        if i is not None and i < len(cols):
            rename[cols[i]] = std
    out.rename(columns=rename, inplace=True)
    return normalize_columns(out)

def normalize_columns(df):
    out = df.copy()

    def clean_money_series(ser: pd.Series) -> pd.Series:
        s = ser.astype(str).fillna("")
        # quitar todo menos dígitos, coma, punto y signo -
        s = s.str.replace(r"[^\d,.\-]", "", regex=True)
        has_dot = s.str.contains(r"\.", regex=True)
        has_comma = s.str.contains(r",", regex=True)
        # ambos presentes -> asumimos formato ES: '.' miles, ',' decimal -> quitar '.' y cambiar ','->'.'
        both = has_dot & has_comma
        s.loc[both] = s.loc[both].str.replace(".", "", regex=False).str.replace(",", ".", regex=False)
        # solo coma -> coma decimal
        only_comma = (~has_dot) & has_comma
        s.loc[only_comma] = s.loc[only_comma].str.replace(",", ".", regex=False)
        # vacíos a 0
        s = s.replace("", "0")
        return pd.to_numeric(s, errors="coerce").fillna(0.0)

    col_aloj = _first_existing(out, ["Nombre alojamiento","Alojamiento","Nombre del alojamiento","Nombre Alojamiento"])
    col_fent = _first_existing(out, ["Fecha entrada","Fecha de entrada"])
    col_fsal = _first_existing(out, ["Fecha salida","Fecha de salida"])
    col_noch = _first_existing(out, ["Noches","noches","Noches ocupadas"])
    col_alq  = _first_existing(out, ["Alquiler con tasas","Ingreso alojamiento","Importe alojamiento"])
    col_ext  = _first_existing(out, ["Ingreso limpieza","Tarifa limpieza","Limpieza","Importe limpieza","Extras con tasas","Gastos de limpieza","Gasto limpieza"])
    col_tot  = _first_existing(out, ["Total reserva con tasas","Total ingresos","Total"])
    col_port = _first_existing(out, ["Web origen","Portal","Canal","Fuente"])
    col_comi = _first_existing(out, ["Comisión Portal/Intermediario: Comisión calculada","Comisión portal","Comisión"])
    col_ivaal= _first_existing(out, ["IVA del alojamiento","IVA alojamiento","IVA del alquiler"])

    rename = {}
    if col_aloj: rename[col_aloj] = "Alojamiento"
    if col_fent: rename[col_fent] = "Fecha entrada"
    if col_fsal: rename[col_fsal] = "Fecha salida"
    if col_noch: rename[col_noch] = "Noches ocupadas"
    if col_alq:  rename[col_alq]  = "Ingreso alojamiento"
    if col_ext:  rename[col_ext]  = "Ingreso limpieza"
    if col_tot:  rename[col_tot]  = "Total ingresos"
    if col_port: rename[col_port] = "Portal"
    if col_comi: rename[col_comi] = "Comisión portal"
    if col_ivaal:rename[col_ivaal]= "IVA del alquiler"

    out.rename(columns=rename, inplace=True)

    # Limpiar y tipar correctamente columnas monetarias (gestiona "1.234,56 €", "1234,56", "1234.56", etc.)
    for c in ["Ingreso alojamiento","Ingreso limpieza","Total ingresos","Comisión portal","IVA del alquiler"]:
        if c in out.columns:
            out[c] = clean_money_series(out[c])

    # Noches y fechas
    if "Noches ocupadas" in out.columns:
        out["Noches ocupadas"] = pd.to_numeric(out["Noches ocupadas"].astype(str).str.replace(r"[^\d\-]", "", regex=True), errors="coerce").fillna(0).round(0).astype(int)
    for c in ["Fecha entrada","Fecha salida"]:
        if c in out.columns:
            out[c] = pd.to_datetime(out[c], errors="coerce", dayfirst=True)

    if "Alojamiento" in out.columns:
        out["Alojamiento"] = out["Alojamiento"].astype(str).str.strip().str.upper()

    return out

# ---------- Carga de reglas desde CSV ----------
def load_rules_csv(path: Optional[str] = None) -> Dict[str, dict]:
    p = Path(path) if path else Path(__file__).resolve().parent / "reglas_apartamentos.csv"
    if not p.exists():
        return {}
    # leer como texto y normalizar
    df = pd.read_csv(p, dtype=str).fillna("")
    df["property_norm"] = df["property"].astype(str).str.strip().str.upper()
    def conv(x):
        try:
            if x == "": return None
            return float(x)
        except Exception:
            return x
    numeric_cols = ["honorarios_pct","honorarios_vat_pct","amenities_amount","cleaning_fee","commission_vat_pct"]
    for col in numeric_cols:
        if col in df.columns:
            df[col] = df[col].apply(conv)
    int_cols = ["honorarios_apply_vat","compute_iva_alquiler","treat_empty_portal_as_booking",
                "skip_booking_vat","split_commission","hon_base_exclude_commission"]
    for col in int_cols:
        if col in df.columns:
            df[col] = df[col].astype(str).replace({"1":"1","0":"0","": "0"}).astype(int)
    rules = df.set_index("property_norm").to_dict(orient="index")
    return rules

RULES_MAP = load_rules_csv(None)

# ---------- Procesador central: usa reglas por piso (CSV) ----------
def process_by_rules(df: pd.DataFrame, rules_map: dict, default_commission_vat: float = 21.0):
    df = normalize_columns(df)
    ensure_required(df, ["Alojamiento","Ingreso alojamiento","Total ingresos","Comisión portal","Portal"], "Procesar por reglas")
    out = df.copy()
    out["Ingreso alojamiento"] = pd.to_numeric(out.get("Ingreso alojamiento",0.0), errors="coerce").fillna(0.0)
    out["Ingreso limpieza"] = pd.to_numeric(out.get("Ingreso limpieza",0.0), errors="coerce").fillna(0.0)
    out["Comisión portal"] = pd.to_numeric(out.get("Comisión portal",0.0), errors="coerce").fillna(0.0)
    out["Total ingresos"] = pd.to_numeric(out.get("Total ingresos",0.0), errors="coerce").fillna(0.0)

    def compute_row(r):
        prop = str(r.get("Alojamiento","")).strip().upper()
        rule = rules_map.get(prop, {})
        ingreso = float(r.get("Ingreso alojamiento",0.0))
        com_orig = float(r.get("Comisión portal",0.0))  # normalmente sin IVA en algunos ficheros
        portal = str(r.get("Portal","") or "").strip().lower()

        # parámetros desde regla (con defaults)
        honorarios_pct = float(rule.get("honorarios_pct") or 0.20)
        honorarios_apply_vat = bool(int(rule.get("honorarios_apply_vat") or 1))
        honorarios_vat_pct = float(rule.get("honorarios_vat_pct") or 21.0)
        amenities_amount = float(rule.get("amenities_amount") or 0.0)
        cleaning_fee = rule.get("cleaning_fee")
        cleaning_fee = float(cleaning_fee) if cleaning_fee not in (None,"") else float(r.get("Ingreso limpieza",0.0))
        compute_iva_alquiler = bool(int(rule.get("compute_iva_alquiler") or 0))
        commission_vat_pct = float(rule.get("commission_vat_pct") if rule.get("commission_vat_pct") not in (None,"") else default_commission_vat)
        treat_empty = bool(int(rule.get("treat_empty_portal_as_booking") or 0))
        skip_booking = bool(int(rule.get("skip_booking_vat") or 0))
        split_comm = bool(int(rule.get("split_commission") or 0))
        hon_base_excl_com = bool(int(rule.get("hon_base_exclude_commission") or 0))

        # ---- aplicar IVA a la comisión (si procede) ----
        is_booking = "booking" in portal
        is_empty = portal == ""
        com_sin_iva = com_orig
        iva_com = 0.0
        # si debemos añadir IVA a la comisión (booking / empty+flag) y no está marcado skip
        if (is_booking or (is_empty and treat_empty)) and commission_vat_pct > 0 and (not skip_booking):
            # com_orig suele venir SIN IVA; añadir IVA opcionalmente
            iva_com = com_orig * (commission_vat_pct / 100.0)
            com_total = com_orig + iva_com
        else:
            com_total = com_orig

        # Si la regla indica que queremos 'split' (desglose) => exponer columnas (sin IVA + IVA) y usar lógica tipo Caso 3
        if split_comm:
            # mantener com_sin_iva como la base original y com_total con IVA añadido si aplica
            # si commission_vat_pct==0 entonces iva_com será 0
            com_sin_iva = com_orig
            iva_com = com_orig * (commission_vat_pct / 100.0) if commission_vat_pct > 0 and not skip_booking else 0.0
            com_total = com_sin_iva + iva_com
        else:
            # si no split pero aplicamos IVA globalmente ya lo hicimos arriba (com_total)
            pass

        # ---- IVA del alquiler si aplica ----
        iva_alq = (ingreso - (ingreso / 1.10)) if compute_iva_alquiler else 0.0

        # ---- base para honorarios ----
        base = ingreso
        if hon_base_excl_com:
            # decidir si excluir com_sin_iva o com_total; mantener la semántica original: excluir comisión SIN IVA
            base = ingreso - com_sin_iva

        # calcular honorarios (con/sin IVA sobre honorarios según regla)
        if honorarios_apply_vat and honorarios_vat_pct:
            honorarios = base * honorarios_pct * (1 + honorarios_vat_pct/100.0)
        else:
            honorarios = base * honorarios_pct

        gasto_limpieza = cleaning_fee
        total_gastos = round(com_total + honorarios + gasto_limpieza + amenities_amount, 2)
        pago_prop = round(float(r.get("Total ingresos",0.0)) - total_gastos, 2)
        pago_recibido = round(float(r.get("Total ingresos",0.0)) - com_total, 2)

        # construir serie resultado, incluir desglose si split_comm
        res = {
            "Comisión portal": round(com_total,2),
            "Honorarios Florit": round(honorarios,2),
            "Gasto limpieza": round(gasto_limpieza,2),
            "Amenities": round(amenities_amount,2),
            "Total Gastos": total_gastos,
            "Pago al propietario": pago_prop,
            "Pago recibido": pago_recibido
        }
        if compute_iva_alquiler:
            res["IVA del alquiler"] = round(iva_alq,2)
        else:
            # dejar None para que el código anterior normalice a 0.0 si procede
            res["IVA del alquiler"] = None

        if split_comm:
            res["Comisión portal (sin IVA)"] = round(com_sin_iva,2)
            res["IVA comisión portal"] = round(iva_com,2)

        return pd.Series(res)

    computed = out.apply(compute_row, axis=1)
    out.update(computed)
    for c in out.columns:
        if c != NIGHTS_COL and pd.api.types.is_numeric_dtype(out[c]):
            out[c] = pd.to_numeric(out[c], errors="coerce").fillna(0.0).round(2)
    return out, 0

# ---------- Export Excel (sin cambios) ----------
BORDER_THIN = Border(left=Side(style="thin"), right=Side(style="thin"),
                     top=Side(style="thin"), bottom=Side(style="thin"))

def write_grouped_sheet(ws, df):
    cols = list(df.columns)
    def write_table(start_row, subdf):
        for j, col in enumerate(cols, start=1):
            cell = ws.cell(row=start_row, column=j, value=col)
            cell.font = Font(bold=True); cell.border = BORDER_THIN
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for i, (_, row) in enumerate(subdf.iterrows(), start=1):
            for j, col in enumerate(cols, start=1):
                val = row[col]; c = ws.cell(row=start_row+i, column=j, value=val); c.border = BORDER_THIN
                if isinstance(val, (int, float)) and not pd.isna(val):
                    if is_nights_col(col): c.number_format = "0"
                    elif is_money_col(col): c.number_format = '#.##0,00" €"'
                    else: c.number_format = "#.##0,00"
                else:
                    c.alignment = Alignment(wrap_text=True)
        sum_row = start_row + len(subdf) + 1
        ws.cell(row=sum_row, column=1, value="TOTAL").font = Font(bold=True); ws.cell(row=sum_row, column=1).border = BORDER_THIN
        for j, col in enumerate(cols, start=1):
            if j == 1: continue
            if pd.api.types.is_numeric_dtype(subdf[col]):
                top = start_row+1; bottom = start_row+len(subdf)
                formula = f"=SUM({get_column_letter(j)}{top}:{get_column_letter(j)}{bottom})"
                c = ws.cell(row=sum_row, column=j, value=formula); c.font = Font(bold=True); c.border = BORDER_THIN
                if is_nights_col(col): c.number_format = "0"
                elif is_money_col(col): c.number_format = '#.##0,00" €"'
                else: c.number_format = "#.##0,00"
            else:
                ws.cell(row=sum_row, column=j, value="").border = BORDER_THIN
        return sum_row + 2
    current_row = 1
    if "Alojamiento" in df.columns:
        for aloj, subdf in df.groupby("Alojamiento"):
            ws.cell(row=current_row, column=1, value=str(aloj)).font = Font(bold=True, size=12)
            current_row += 1
            current_row = write_table(current_row, subdf)
    else:
        current_row = write_table(current_row, df)
    for j, col in enumerate(cols, start=1):
        max_len = len(str(col))
        for r in range(1, ws.max_row+1):
            v = ws.cell(row=r, column=j).value
            if v is not None: max_len = max(max_len, len(str(v)))
        ws.column_dimensions[get_column_letter(j)].width = min(max_len+2, 45)

def build_excel_single(df_final, filename="Liquidacion.xlsx"):
    wb = Workbook(); ws = wb.active; ws.title = "Liquidación"
    write_grouped_sheet(ws, df_final)
    bio = BytesIO(); wb.save(bio); bio.seek(0)
    st.download_button("📥 Descargar Excel (Liquidación)", bio.getvalue(),
                       file_name=filename,
                       mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

# ---------- UI ----------
st.title("📊 LIQUIDACIONES (por reglas CSV)")
st.caption("Genera liquidaciones usando reglas por piso desde reglas_apartamentos.csv")

with st.sidebar:
    st.header("Parámetros")
    c1, c2 = st.columns(2)
    with c1:
        start_date = st.date_input("Desde", value=date(date.today().year, date.today().month, 1))
    with c2:
        end_date = st.date_input("Hasta", value=date(date.today().year, date.today().month, 28))
    st.divider()
    st.checkbox("Lectura por letras (fallback)", value=False, key="by_letters")
    use_rules = st.checkbox("Usar reglas CSV (reglas_apartamentos.csv) - calcular por piso", value=True, key="use_rules_csv")
    st.divider()
    # filtro por piso (lista desde el CSV de reglas)
    pisos_options = sorted(RULES_MAP.keys()) if RULES_MAP else []
    st.multiselect("Filtrar por piso (Alojamiento)", options=pisos_options, key="filter_pisos")
    header_second_row = st.checkbox("La cabecera está en la segunda fila (leer desde la fila 2)", value=False)
    generate = st.button("Generar liquidación")

file = st.file_uploader("Sube el archivo de reservas (.xlsx)", type=["xlsx"], key="reservas_upl")

def normalize_liq_for_period(df_norm, start_date, end_date):
    if "Fecha entrada" in df_norm.columns:
        mask = (df_norm["Fecha entrada"] >= pd.to_datetime(start_date)) & (df_norm["Fecha entrada"] <= pd.to_datetime(end_date))
        df_norm = df_norm[mask]
    return df_norm

if generate:
    if not file:
        st.error("Sube primero el archivo de reservas (.xlsx)."); st.stop()
    if not use_rules:
        st.error("La aplicación ahora exige 'Usar reglas CSV'. Activa la opción y vuelve a generar."); st.stop()

    header_row = 1 if header_second_row else 0
    df_in = pd.read_excel(file, header=header_row)
    df_in = ensure_unique_columns(df_in)
    df_norm = normalize_columns_by_letters(df_in) if st.session_state.by_letters else normalize_columns(df_in)
    df_norm = ensure_unique_columns(df_norm)
    df_norm = normalize_liq_for_period(df_norm, start_date, end_date)

    # aplicar filtro por piso si hay selección
    selected_pisos = st.session_state.get("filter_pisos", []) or []
    if selected_pisos:
        selected_norm = [s.strip().upper() for s in selected_pisos]
        df_norm = df_norm[df_norm["Alojamiento"].isin(selected_norm)]
        if df_norm.empty:
            st.warning("No hay reservas para los pisos seleccionados en el período.")
            st.stop()

    if RULES_MAP == {}:
        st.error("No se encontró reglas_apartamentos.csv o está vacío en la carpeta del script."); st.stop()

    df_out, warn = process_by_rules(df_norm, RULES_MAP)
    sort_cols = [c for c in ["Alojamiento","Fecha entrada"] if c in df_out.columns]
    if sort_cols: df_out = df_out.sort_values(by=sort_cols)
    st.success(f"Liquidación generada (Por reglas CSV) • {start_date.strftime('%d/%m/%Y')}–{end_date.strftime('%d/%m/%Y')}")
    show_table_es_grouped(df_out, "Tabla de liquidaciones (por reglas)")
    file_case_name = f"Liquidacion_REGLAS_CSV_{start_date.strftime('%Y%m%d')}_{end_date.strftime('%Y%m%d')}.xlsx"
    build_excel_single(df_out, filename=file_case_name)
    st.session_state["df_liq_all"] = df_out.copy()
    st.session_state["df_liq_label"] = "Por reglas CSV"