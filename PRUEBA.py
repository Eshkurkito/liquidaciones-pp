import streamlit as st
import pandas as pd
import numpy as np
from datetime import date
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font

st.set_page_config(page_title="Liquidaciones dinámicas", page_icon="📊", layout="wide")

# =====================================================
# CARGA REGLAS CSV (AUTODETECTA SEPARADOR)
# =====================================================

#@st.cache_data
def load_reglas():
    try:
        reglas = pd.read_csv("reglas_apartamentos.csv", sep=None, engine="python")
    except FileNotFoundError:
        st.error("No se encuentra reglas_apartamentos.csv en el repositorio.")
        st.stop()

    reglas.columns = reglas.columns.str.strip()

    if "Property" not in reglas.columns:
        st.error(f"Columnas detectadas en CSV: {list(reglas.columns)}")
        st.error("No existe columna 'Property' en reglas_apartamentos.csv")
        st.stop()

    reglas["Property"] = reglas["Property"].astype(str).str.strip().str.upper()

    bool_cols = [
    "honorarios_apply_vat",
    "compute_iva_alquiler",
    "treat_empty_portal_as_booking",
    "skip_booking_vat",
    "hon_base_exclude_commission",
    "hon_base_use_commission_without_vat",
    "hon_base_exclude_rent_vat",
    "include_rent_vat_in_expenses"
    ]

    for col in bool_cols:
        if col in reglas.columns:
            reglas[col] = (
                pd.to_numeric(reglas[col], errors="coerce")
                .fillna(0)
                .astype(int)
                .map({1: True, 0: False, 2: False})
            )

            
    if "self_managed" in reglas.columns:
        reglas["self_managed"] = (
        pd.to_numeric(reglas["self_managed"], errors="coerce")
        .fillna(0)
        .astype(int)
    )

    # Convertir commission_vat_pct a numérico
    if "commission_vat_pct" in reglas.columns:
        reglas["commission_vat_pct"] = pd.to_numeric(
            reglas["commission_vat_pct"], errors="coerce"
        ).fillna(0)

    # Nuevo: rent_vat_pct (IVA aplicado al alquiler). Si no existe, default 10%
    if "rent_vat_pct" in reglas.columns:
        reglas["rent_vat_pct"] = pd.to_numeric(
            reglas["rent_vat_pct"], errors="coerce"
        ).fillna(10.0)
    else:
        reglas["rent_vat_pct"] = 10.0

    # Corregir filas con compute_iva_alquiler True pero rent_vat_pct == 0 → asumir 10%
    mask_fix = (reglas.get("compute_iva_alquiler", False) == 1) & (reglas["rent_vat_pct"] <= 0)
    if mask_fix.any():
        reglas.loc[mask_fix, "rent_vat_pct"] = 10.0

    return reglas


# =====================================================
# NORMALIZACIÓN ROBUSTA (COPIADA DEL BACKUP)
# =====================================================

def normalize_columns(df):

    out = df.copy()

    def first_existing(candidates):
        norm_map = {str(c).strip().lower(): c for c in out.columns}
        for cand in candidates:
            key = cand.strip().lower()
            if key in norm_map:
                return norm_map[key]
        return None

    col_aloj = first_existing(["Nombre alojamiento","Alojamiento","Nombre del alojamiento"])
    col_fent = first_existing(["Fecha entrada","Fecha de entrada"])
    col_fsal = first_existing(["Fecha salida","Fecha de salida"])
    col_noch = first_existing(["Noches","Noches ocupadas"])
    col_alq  = first_existing(["Alquiler con tasas","Ingreso alojamiento","Importe alojamiento"])
    col_ext  = first_existing([
        "Ingreso limpieza","Tarifa limpieza","Limpieza","Importe limpieza",
        "Extras con tasas","Gastos de limpieza","Gasto limpieza"
    ])
    col_tot  = first_existing(["Total reserva con tasas","Total ingresos","Total"])
    col_port = first_existing(["Web origen","Portal","Canal"])
    col_comi = first_existing([
        "Comisión Portal/Intermediario: Comisión calculada",
        "Comisión portal","Comisión"
    ])
    
    # Buscar columna comisión de forma flexible
    # Detectar columna de comisión de forma flexible
    col_comi = None

    for col in out.columns:
        nombre = str(col).lower()
        if "comisión" in nombre or "comision" in nombre:
            col_comi = col
            break

    if col_comi:
        out.rename(columns={col_comi: "Comisión portal"}, inplace=True)



    rename = {}
    if col_aloj: rename[col_aloj] = "Alojamiento"
    if col_fent: rename[col_fent] = "Fecha entrada"
    if col_fsal: rename[col_fsal] = "Fecha salida"
    if col_noch: rename[col_noch] = "Noches ocupadas"
    if col_alq:  rename[col_alq]  = "Ingreso alojamiento"
    if col_ext:  rename[col_ext]  = "Ingreso limpieza"
    if col_tot:  rename[col_tot]  = "Total ingresos"
    if col_port: rename[col_port] = "Portal"
    if col_comi: 
        rename[col_comi] = "Comisión portal"

    out.rename(columns=rename, inplace=True)

    if "Ingreso limpieza" not in out.columns:
        out["Ingreso limpieza"] = 0.0

    for c in ["Ingreso alojamiento","Ingreso limpieza","Total ingresos","Comisión portal"]:
        if c in out.columns:
            out[c] = pd.to_numeric(out[c], errors="coerce").fillna(0.0)

    for c in ["Fecha entrada","Fecha salida"]:
        if c in out.columns:
            out[c] = pd.to_datetime(out[c], errors="coerce", dayfirst=True)

    if "Alojamiento" in out.columns:
        out["Alojamiento"] = out["Alojamiento"].astype(str).str.strip().str.upper()
        
        
    # -------------------------
    # HUESPEDES TOTALES
    # -------------------------

    col_adultos = first_existing(["Adultos","adultos"])
    col_ninos   = first_existing(["Niños","Ninos","niños","ninos"])

    if col_adultos:
        out["Adultos"] = pd.to_numeric(out[col_adultos], errors="coerce").fillna(0)
    else:
        out["Adultos"] = 0

    if col_ninos:
        out["Niños"] = pd.to_numeric(out[col_ninos], errors="coerce").fillna(0)
    else:
        out["Niños"] = 0

    out["Huéspedes totales"] = out["Adultos"] + out["Niños"]

    # Garantizar que exista Comisión portal
    if "Comisión portal" not in out.columns:
        out["Comisión portal"] = 0.0

    return out


# =====================================================
# MOTOR DE LIQUIDACIÓN DINÁMICO
# =====================================================

def process_dynamic(df, reglas):


    df = df.merge(reglas, left_on="Alojamiento", right_on="Property", how="left")
    df = df.loc[:, ~df.columns.duplicated()]

    # Diagnostics: listar alojamientos del Excel vs reglas
    alojamientos_en_excel = sorted(df["Alojamiento"].dropna().astype(str).unique())
    reglas_props = sorted(reglas["Property"].dropna().astype(str).unique())
    unmatched = [a for a in alojamientos_en_excel if a not in reglas_props]

    # 🔥 SOLO apartamentos con reglas
    df_matched = df[df["Property"].notna()].copy()

    # Columnas mínimas que esperan el resto de la app (evita KeyError si no hay filas)
    columnas_minimas = [
        "Alojamiento",
        "Fecha entrada",
        "Fecha salida",
        "Noches ocupadas",
        "Huéspedes totales",
        "Ingreso alojamiento",
        "IVA del alquiler",
        "Ingreso limpieza",
        "Total ingresos",
        "Portal",
        "Comisión portal",
        "Honorarios Florit",
        "Gasto limpieza",
        "Amenities",
        "Total Gastos",
        "Pago al propietario",
        "Pago recibido",
    ]

    
    if df_matched.empty:
        st.warning("No hay apartamentos del CSV en el período seleccionado.")
        if unmatched:
            st.info(
                "Alojamientos en reservas no encontrados en reglas: "
                + ", ".join(unmatched[:50])
            )
        else:
            st.info(
                "Ningún alojamiento de las reservas coincide con los de reglas. "
                "Revisa formatos (espacios, mayúsculas, acentos). Ejemplos en reglas: "
                + ", ".join(reglas_props[:50])
            )
        # Devolver DataFrame vacío con columnas mínimas (evita KeyError en sumas/filtrados)
        return pd.DataFrame(columns=columnas_minimas)

    df = df_matched
    # -------------------------
    # COMISIÓN PORTAL
    # -------------------------
   
    # Asegurar que es numérica
    df["Comisión portal"] = pd.to_numeric(
        df["Comisión portal"], errors="coerce"
    ).fillna(0)
    
    mask_muchosol = df["Portal"].str.upper().str.contains("MUCHOSOL", na=False)

    df.loc[mask_muchosol, "Comisión portal"] = (
        df.loc[mask_muchosol, "Ingreso alojamiento"] * 0.28
    )


    # Comisión SIN IVA (viene del Excel)
    df["Comisión portal sin IVA"] = df["Comisión portal"].copy()
    

    # 🔥 CALCULAR IVA comisión usando commission_vat_pct del CSV
    df["IVA comisión portal"] = (
        df["Comisión portal sin IVA"] * df["commission_vat_pct"] / 100
    )

    # Comisión CON IVA real
    df["Comisión portal con IVA"] = (
        df["Comisión portal sin IVA"] + df["IVA comisión portal"]
    )
    
    # Mostrar comisión según IVA configurado
    df["Comisión portal visible"] = np.where(
        df["commission_vat_pct"] > 0,
        df["Comisión portal con IVA"],
        df["Comisión portal sin IVA"]
    )


    # -- quitar debug directo --
    # st.write("DEBUG Comisión SIN IVA:", df["Comisión portal sin IVA"].iloc[0])
    # st.write("DEBUG commission_vat_pct:", df["commission_vat_pct"].iloc[0])
    # st.write("DEBUG IVA comisión:", df["IVA comisión portal"].iloc[0])
    # st.write("DEBUG Comisión CON IVA:", df["Comisión portal con IVA"].iloc[0])

    
    # -------------------------
    # IVA ALQUILER
    # -------------------------

    df["IVA del alquiler"] = 0.0

    mask_iva = df["compute_iva_alquiler"] == True

    # Usar rent_vat_pct por fila (1 + pct/100). Evita constante 1.10
    if "rent_vat_pct" not in df.columns:
        df["rent_vat_pct"] = 10.0

    df.loc[mask_iva, "IVA del alquiler"] = (
        df.loc[mask_iva, "Ingreso alojamiento"]
        - (df.loc[mask_iva, "Ingreso alojamiento"] / (1 + df.loc[mask_iva, "rent_vat_pct"] / 100))
    )

    # -------------------------
    # HONORARIOS (VERSIÓN RÁPIDA)
    # -------------------------

    base = df["Ingreso alojamiento"].copy()

    # Excluir comisión si aplica (usar .loc para preservar índice)
    mask_exclude = df["hon_base_exclude_commission"] == True
    mask_without_vat = df["hon_base_use_commission_without_vat"] == True

    # Si usa comisión SIN IVA
    mask = mask_exclude & mask_without_vat
    base.loc[mask] = base.loc[mask] - df.loc[mask, "Comisión portal sin IVA"]

    # Si usa comisión CON IVA
    mask = mask_exclude & (~mask_without_vat)
    base.loc[mask] = base.loc[mask] - df.loc[mask, "Comisión portal con IVA"]

    # Excluir IVA si aplica
    if "hon_base_exclude_rent_vat" not in df.columns:
        df["hon_base_exclude_rent_vat"] = False

    mask_iva_base = df["hon_base_exclude_rent_vat"] == True
    base.loc[mask_iva_base] = base.loc[mask_iva_base] - df.loc[mask_iva_base, "IVA del alquiler"]

    df["Honorarios Florit"] = base * df["honorarios_pct"]

    # Aplicar IVA a honorarios si corresponde
    mask_vat = df["honorarios_apply_vat"] == True
    df.loc[mask_vat, "Honorarios Florit"] *= (
        1 + df.loc[mask_vat, "honorarios_vat_pct"] / 100
    )


    # -------------------------
    # LIMPIEZA
    # -------------------------

    df["Gasto limpieza"] = np.where(
        df["cleaning_fee"] > 0,
        df["cleaning_fee"],
        df["Ingreso limpieza"]
    )

    # -------------------------
    # AMENITIES
    # -------------------------

    df["Amenities"] = df["amenities_amount"].fillna(0)

    # -------------------------
    # TOTALES
    # -------------------------

    # Incluir IVA del alquiler sólo si la regla marca include_rent_vat_in_expenses == True
    iva_alq = df.get("IVA del alquiler", 0).fillna(0)
    iva_flag = df.get("include_rent_vat_in_expenses", False).fillna(0)
    # asegúrate de que iva_flag sea 0/1
    if iva_flag.dtype == bool:
        iva_flag = iva_flag.astype(int)
    else:
        iva_flag = pd.to_numeric(iva_flag, errors="coerce").fillna(0).astype(int)
    
    # ---------------------------------
    # IVA BOOKING (según skip_booking_vat)
    # ---------------------------------

    df["IVA Booking"] = np.where(
        df["Portal"].str.contains("BOOKING", case=False, na=False),
        df["Comisión portal sin IVA"] * 0.21,
        0
    )

    skip_flag = df.get("skip_booking_vat", False)

    if skip_flag.dtype == bool:
        skip_flag = skip_flag.astype(int)
    else:
        skip_flag = pd.to_numeric(skip_flag, errors="coerce").fillna(0).astype(int)

    # Si skip = 0 → sumar IVA
    # Si skip = 1 → no sumar
    booking_vat_final = df["IVA Booking"] * (1 - skip_flag)

    
    df["Total Gastos"] = (
        df["Comisión portal con IVA"]
        + booking_vat_final
        + df["Honorarios Florit"]
        + df["Gasto limpieza"]
        + df["Amenities"]
        + (iva_alq * iva_flag)
        
    )
    
    df["Pago al propietario"] = df["Total ingresos"] - df["Total Gastos"]
    
    # ---------------------------------
    # SELF MANAGED (Florit = propietario)
    # ---------------------------------

    
    df["Pago recibido"] = df["Total ingresos"] - df["Comisión portal"]

    columnas_finales = [
        "Alojamiento",
        "Fecha entrada",
        "Fecha salida",
        "Noches ocupadas",
        "Huéspedes totales",
        "Ingreso alojamiento",
        "IVA del alquiler",
        "Ingreso limpieza",
        "Total ingresos",
        "Portal",
        "Comisión portal visible",
        "Honorarios Florit",
        "Gasto limpieza",
        "Amenities",
        "Total Gastos",
        "Pago al propietario",
        "Pago recibido",
    ]

    columnas_existentes = [c for c in columnas_finales if c in df.columns]
    
    df = df.sort_values(by=["Alojamiento", "Fecha entrada"])
    
    if "self_managed" in df.columns:

        mask_self = df["self_managed"] == 1

        # Asegurar que todo es numérico
        cols_needed = [
            "Total ingresos",
            "Comisión portal",
            "Gasto limpieza",
            "Amenities"
        ]

        for col in cols_needed:
            if col in df.columns:
                df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0)

        # Honorarios = beneficio real
        df.loc[mask_self, "Honorarios Florit"] = (
            df.loc[mask_self, "Total ingresos"]
            - df.loc[mask_self, "Comisión portal"]
            - df.loc[mask_self, "Gasto limpieza"]
            - df.loc[mask_self, "Amenities"]
        )

        # Total Gastos = solo gastos reales (sin honorarios)
        df.loc[mask_self, "Total Gastos"] = (
            df.loc[mask_self, "Comisión portal"]
            + df.loc[mask_self, "Gasto limpieza"]
            + df.loc[mask_self, "Amenities"]
        )

        # Pago propietario = mismo importe que honorarios
        df.loc[mask_self, "Pago al propietario"] = df.loc[mask_self, "Honorarios Florit"]
        
    return df[columnas_existentes]


# =====================================================
# EXPORTAR EXCEL
# =====================================================

def build_excel(df):

    wb = Workbook()

    # Separar propios y terceros
    df_propios = df[df.get("self_managed", 0) == 1]
    df_terceros = df[df.get("self_managed", 0) != 1]

    def escribir_hoja(ws, dataframe):

        row_cursor = 1

        for aloj, subdf in dataframe.groupby("Alojamiento"):

            ws.cell(row=row_cursor, column=1, value=aloj).font = Font(bold=True)
            row_cursor += 1

            cols = list(subdf.columns)

            # Cabecera
            for col_idx, col in enumerate(cols, 1):
                ws.cell(row=row_cursor, column=col_idx, value=col).font = Font(bold=True)

            row_cursor += 1

            # Reservas
            
            for _, row in subdf.iterrows():
                for col_idx, col in enumerate(cols, 1):

                    value = row[col]
                    cell = ws.cell(row=row_cursor, column=col_idx)

                     # Fechas sin hora
                    if isinstance(value, pd.Timestamp):
                        cell.value = value.date()

                    # Formato moneda
                    elif isinstance(value, (int, float)) and col in [
                        "Ingreso alojamiento",
                        "IVA del alquiler",
                        "Ingreso limpieza",
                        "Total ingresos",
                        "Comisión portal visible",
                        "Honorarios Florit",
                        "Gasto limpieza",
                        "Amenities",
                        "Total Gastos",
                        "Pago al propietario",
                        "Pago recibido"
                    ]:
                        cell.value = value
                        cell.number_format = '#,##0.00 €'

                    else:
                        cell.value = value

                row_cursor += 1
            # Fila TOTAL
            for col_idx, col in enumerate(cols, 1):
                if pd.api.types.is_numeric_dtype(subdf[col]):
                    total_value = subdf[col].sum()
                    ws.cell(row=row_cursor, column=col_idx, value=total_value).font = Font(bold=True)
                else:
                    if col == "Fecha entrada":
                        ws.cell(row=row_cursor, column=col_idx, value="TOTAL").font = Font(bold=True)

            row_cursor += 2

    # Hoja Propios
    if not df_propios.empty:
        ws1 = wb.active
        ws1.title = "Propios"
        escribir_hoja(ws1, df_propios)
    else:
        ws1 = wb.active
        ws1.title = "Propios"

    # Hoja Terceros
    if not df_terceros.empty:
        ws2 = wb.create_sheet(title="Terceros")
        escribir_hoja(ws2, df_terceros)

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)

    return bio



st.title("📊 Liquidaciones dinámicas")

# 🔹 TABS DESPUÉS


# INPUTS
start_date = st.date_input("Desde", key="desde_global")
end_date = st.date_input("Hasta", key="hasta_global")
file_reservas = st.file_uploader("Sube archivo (.xlsx)", type=["xlsx"], key="file_global")
df_base = None

if file_reservas:
    df_base = pd.read_excel(file_reservas)
    df_base = normalize_columns(df_base)


st.divider()

tab1, tab2 = st.tabs(["Liquidaciones", "Previsión Tesorería"])

reglas = load_reglas()

if "df_final" not in st.session_state:
    st.session_state.df_final = None
    
if "df_tesoreria" not in st.session_state:
    st.session_state.df_tesoreria = None


st.divider()  # Opcional para separar visualmente



with tab1:
    
    

    # ---------------------------------
    # CARGAR REGLAS Y CREAR FILTRO
    # ---------------------------------


    alojamientos_gestionados = sorted(reglas["Property"].unique())

    # Inicializar estado si no existe
    if "alojamientos_sel" not in st.session_state:
        st.session_state.alojamientos_sel = alojamientos_gestionados

    col1, col2 = st.columns(2)

    with col1:
        if st.button("Seleccionar todos"):
            st.session_state.alojamientos_sel = alojamientos_gestionados

    with col2:
        if st.button("Quitar todos"):
            st.session_state.alojamientos_sel = []

    alojamientos_seleccionados = st.multiselect(
    "Selecciona alojamiento(s)",
    options=alojamientos_gestionados,
    default=st.session_state.alojamientos_sel,
    key="alojamientos_sel"
    )


    # ---------------------------------
    # BOTÓN
    # ---------------------------------

    if st.button("Generar liquidación"):

        if not file_reservas:
            st.error("Sube el archivo de reservas.")
            st.stop()

        df_res = df_base.copy()


        # 🔥 FILTRAR ANTES DEL CÁLCULO
        df_res = df_res[df_res["Alojamiento"].isin(alojamientos_seleccionados)]

        df_res = df_res[
            (df_res["Fecha entrada"] >= pd.to_datetime(start_date))
            & (df_res["Fecha entrada"] <= pd.to_datetime(end_date))
        ]

        if df_res.empty:
            st.warning("No hay reservas para los alojamientos seleccionados.")
            st.stop()

        st.session_state.df_final = process_dynamic(df_res, reglas)

        if st.session_state.df_final is None or st.session_state.df_final.empty:
            st.warning("No se ha generado ninguna liquidación tras aplicar las reglas. Revisa los alojamientos y el periodo.")
            st.stop()

        # Mostrar tablas por apartamento y botón de descarga
        df_show = st.session_state.df_final.copy()
        if not df_show.empty:
            # Botón descarga Excel
            bio = build_excel(df_show)
            st.download_button(
                "📥 Descargar liquidaciones (Excel)",
                data=bio,
                file_name="liquidaciones.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

            # Expanders por apartamento con su tabla
            for aloj, sub in df_show.groupby("Alojamiento"):
                total_hon = sub["Honorarios Florit"].sum() if "Honorarios Florit" in sub.columns else 0.0
                with st.expander(f"{aloj} — Honorarios: {total_hon:,.2f} €", expanded=False):
                    # Preparar fila TOTAL con sumas para columnas numéricas
                    sub_display = sub.reset_index(drop=True).copy()
                    
                    # ---------------------------------
                    # FORMATO VISUAL
                    # ---------------------------------

                    # Quitar hora en fechas
                    if "Fecha entrada" in sub_display.columns:
                        sub_display["Fecha entrada"] = pd.to_datetime(
                             sub_display["Fecha entrada"]
                        ).dt.date

                    if "Fecha salida" in sub_display.columns:
                        sub_display["Fecha salida"] = pd.to_datetime(
                            sub_display["Fecha salida"]
                        ).dt.date

                    # Columnas formato €
                    columnas_euro = [
                        "Ingreso alojamiento",
                        "IVA del alquiler",
                        "Ingreso limpieza",
                        "Total ingresos",
                        "Comisión portal visible",
                        "Honorarios Florit",
                        "Gasto limpieza",
                        "Amenities",
                        "Total Gastos",
                        "Pago al propietario",
                        "Pago recibido"
                    ]

                    formato = {
                        col: "{:,.2f} €"
                        for col in columnas_euro
                        if col in sub_display.columns
                    }
                    
                    numeric_cols = sub_display.select_dtypes(include=[np.number]).columns.tolist()
                    totals_row = {c: sub_display[c].sum() for c in numeric_cols}
                    # Rellenar resto de columnas con cadena vacía y marcar TOTAL en Fecha entrada si existe
                    for c in sub_display.columns:
                        if c not in totals_row:
                            totals_row[c] = ""
                    if "Fecha entrada" in sub_display.columns:
                        totals_row["Fecha entrada"] = "TOTAL"
                    # Añadir fila TOTAL al final
                    sub_display = pd.concat([sub_display, pd.DataFrame([totals_row])], ignore_index=True)
                    st.dataframe(
                        sub_display.style.format(formato),
                        use_container_width=True
                )
# =====================================================
# TAB 2 → PREVISIÓN TESORERÍA
# =====================================================

with tab2:

    st.header("📈 Previsión Tesorería – Honorarios Florit")

    if df_base is None:
        st.info("Sube archivo para ver previsión.")
        st.stop()

    # 🔹 FORMULARIO
    with st.form("form_tesoreria"):

        alojamientos_disponibles = sorted(reglas["Property"].unique())

        # Inicializar estado si no existe
        if "alojamientos_tab2" not in st.session_state:
            st.session_state.alojamientos_tab2 = alojamientos_disponibles

        col1, col2 = st.columns(2)

        with col1:
            if st.form_submit_button("Seleccionar todos"):
                st.session_state.alojamientos_tab2 = alojamientos_disponibles

        with col2:
            if st.form_submit_button("Quitar todos"):
                st.session_state.alojamientos_tab2 = []

        alojamientos_tab2 = st.multiselect(
            "Filtrar alojamiento(s)",
            options=alojamientos_disponibles,
            default=st.session_state.alojamientos_tab2,
            key="alojamientos_tab2"
        )

        fecha_corte = st.date_input(
            "Fecha de corte",
            value=end_date,
            key="fecha_corte_tesoreria"
        )

        calcular_tesoreria = st.form_submit_button("Calcular previsión")


    # 🔹 SOLO SI SE PULSA BOTÓN
    if calcular_tesoreria:

        df_periodo = df_base.copy()

        if alojamientos_tab2:
            df_periodo = df_periodo[
                df_periodo["Alojamiento"].isin(alojamientos_tab2)
            ]

        df_periodo = df_periodo[
            (df_periodo["Fecha entrada"] >= pd.to_datetime(start_date)) &
            (df_periodo["Fecha entrada"] <= pd.to_datetime(end_date))
        ]

        df_periodo = process_dynamic(df_periodo, reglas)

        if df_periodo is None or df_periodo.empty:
            st.warning("No hay datos para la previsión tras aplicar las reglas. Revisa alojamientos y periodo.")
        else:
            st.session_state.df_tesoreria = df_periodo

    # 🔹 MOSTRAR SOLO SI YA HAY RESULTADO
    if st.session_state.df_tesoreria is not None:

        df_periodo = st.session_state.df_tesoreria

        df_periodo["Año"] = df_periodo["Fecha entrada"].dt.year
        df_periodo["Mes"] = df_periodo["Fecha entrada"].dt.month
        df_periodo["Mes_nombre"] = df_periodo["Fecha entrada"].dt.strftime("%b")


        df_corte = df_periodo[
            df_periodo["Fecha entrada"] <= pd.to_datetime(
                st.session_state.fecha_corte_tesoreria
            )
        ]

        honorarios_corte = df_corte["Honorarios Florit"].sum()
        honorarios_periodo = df_periodo["Honorarios Florit"].sum()
        

        # =====================================================
        # COMPARACIÓN YTD VS AÑO ANTERIOR
        # =====================================================

        anio_actual = end_date.year
        anio_anterior = anio_actual - 1

        fecha_inicio_anterior = st.session_state.desde_global.replace(year=anio_anterior)
        fecha_fin_anterior = st.session_state.hasta_global.replace(year=anio_anterior)

        df_anterior = df_base.copy()

        if st.session_state.alojamientos_tab2:
            df_anterior = df_anterior[
                df_anterior["Alojamiento"].isin(st.session_state.alojamientos_tab2)
            ]

        df_anterior = df_anterior[
            (df_anterior["Fecha entrada"] >= pd.to_datetime(fecha_inicio_anterior)) &
            (df_anterior["Fecha entrada"] <= pd.to_datetime(fecha_fin_anterior))
        ]

        df_anterior = process_dynamic(df_anterior, reglas)

        # Asegurar que no intentamos sumar una columna inexistente
        if df_anterior is None or df_anterior.empty or "Honorarios Florit" not in df_anterior.columns:
            hon_anterior = 0.0
        else:
            hon_anterior = df_anterior["Honorarios Florit"].sum()

        hon_actual = honorarios_periodo
        hon_anterior = hon_anterior

        variacion_pct = (
            (hon_actual - hon_anterior) / hon_anterior * 100
            if hon_anterior > 0 else 0
        )

        porcentaje = (
            honorarios_corte / honorarios_periodo * 100
            if honorarios_periodo > 0 else 0
        )

        pendiente = honorarios_periodo - honorarios_corte

        col1, col2, col3, col4 = st.columns(4)

        col1.metric("💶 Generado hasta corte", f"{honorarios_corte:,.2f} €")
        col2.metric("📅 Total periodo", f"{honorarios_periodo:,.2f} €")
        col3.metric("📊 % ejecutado", f"{porcentaje:,.1f} %")
        col4.metric("🔮 Pendiente periodo", f"{pendiente:,.2f} €")

        st.divider()

        col5, col6, col7 = st.columns(3)

        col5.metric(f"💶 {anio_actual}", f"{hon_actual:,.2f} €")
        col6.metric(f"💶 {anio_anterior}", f"{hon_anterior:,.2f} €")
        col7.metric("📈 Variación %", f"{variacion_pct:,.1f} %")

        st.divider()

        ranking = (
            df_periodo
            .groupby("Alojamiento")["Honorarios Florit"]
            .sum()
            .sort_values(ascending=False)
            .reset_index()
        )

        st.divider()
        st.subheader("📅 Honorarios mensuales por apartamento")

        # ---------------------------------
        # HONORARIOS MENSUALES SEPARADOS
        # ---------------------------------

        df_periodo["Año"] = df_periodo["Fecha entrada"].dt.year
        df_periodo["Mes"] = df_periodo["Fecha entrada"].dt.month

        # Separar propios y terceros
        df_propios = df_periodo[df_periodo["self_managed"] == 1]
        df_terceros = df_periodo[df_periodo["self_managed"] != 1]

        def construir_tabla_mensual(df_input):

            if df_input.empty:
                return pd.DataFrame()

            tabla = (
                df_input
                .groupby(["Alojamiento", "Mes"])["Honorarios Florit"]
                .sum()
                .reset_index()
            )

            pivot = (
                tabla
                .pivot(index="Alojamiento", columns="Mes", values="Honorarios Florit")
                .fillna(0)
            )

            # Orden enero → diciembre
            orden_meses = list(range(1, 13))
            pivot = pivot.reindex(columns=orden_meses, fill_value=0)

            nombres_meses = {
                1: "Ene", 2: "Feb", 3: "Mar", 4: "Abr",
                5: "May", 6: "Jun", 7: "Jul", 8: "Ago",
                9: "Sep", 10: "Oct", 11: "Nov", 12: "Dic"
            }

            pivot.rename(columns=nombres_meses, inplace=True)
            pivot["TOTAL AÑO"] = pivot.sum(axis=1)

            return pivot.round(2)

        pivot_propios = construir_tabla_mensual(df_propios)
        pivot_terceros = construir_tabla_mensual(df_terceros)
        
        # =====================================================
        # KPI MIX NEGOCIO
        # =====================================================

        total_propios = df_propios["Honorarios Florit"].sum()
        total_terceros = df_terceros["Honorarios Florit"].sum()
        total_global = total_propios + total_terceros

        mix_propios = (total_propios / total_global * 100) if total_global > 0 else 0
        mix_terceros = (total_terceros / total_global * 100) if total_global > 0 else 0

        st.divider()
        st.subheader("📊 Mix de negocio")

        col1, col2, col3, col4 = st.columns(4)

        col1.metric("🏠 Total Propios", f"{total_propios:,.2f} €")
        col2.metric("🤝 Total Terceros", f"{total_terceros:,.2f} €")
        col3.metric("📊 % Propios", f"{mix_propios:,.1f} %")
        col4.metric("📊 % Terceros", f"{mix_terceros:,.1f} %")
        
        # =====================================================
        # GRÁFICO COMPARATIVO MENSUAL
        # =====================================================

        import plotly.express as px

        # Agrupar por mes
        mensual_propios = (
            df_propios.groupby("Mes")["Honorarios Florit"]
            .sum()
            .reindex(range(1,13), fill_value=0)
        )

        mensual_terceros = (
            df_terceros.groupby("Mes")["Honorarios Florit"]
            .sum()
            .reindex(range(1,13), fill_value=0)
        )

        df_grafico = pd.DataFrame({
            "Mes": range(1,13),
            "Propios": mensual_propios.values,
            "Terceros": mensual_terceros.values
        })

        nombres_meses = {
            1:"Ene",2:"Feb",3:"Mar",4:"Abr",5:"May",6:"Jun",
            7:"Jul",8:"Ago",9:"Sep",10:"Oct",11:"Nov",12:"Dic"
        }

        df_grafico["Mes"] = df_grafico["Mes"].map(nombres_meses)

        fig = px.bar(
            df_grafico,
            x="Mes",
            y=["Propios","Terceros"],
            barmode="group",
            title="Comparativa mensual Honorarios"
        )

        st.plotly_chart(fig, use_container_width=True)

        # Mostrar tablas
        if not pivot_propios.empty:
            st.subheader("🏠 Alojamientos propios")
            st.dataframe(pivot_propios, use_container_width=True)

        if not pivot_terceros.empty:
            st.subheader("🤝 Alojamientos de terceros")
            st.dataframe(pivot_terceros, use_container_width=True)

        def build_excel_mensual(propios_df, terceros_df):

            wb = Workbook()

            if not propios_df.empty:
                ws1 = wb.active
                ws1.title = "Propios"

                for col_idx, col in enumerate(propios_df.columns.insert(0, "Alojamiento"), 1):
                    ws1.cell(row=1, column=col_idx, value=col).font = Font(bold=True)

                for row_idx, (index, row) in enumerate(propios_df.iterrows(), 2):
                    ws1.cell(row=row_idx, column=1, value=index)
                    for col_idx, value in enumerate(row, 2):
                        ws1.cell(row=row_idx, column=col_idx, value=value)

            if not terceros_df.empty:
                ws2 = wb.create_sheet(title="Terceros")

                for col_idx, col in enumerate(terceros_df.columns.insert(0, "Alojamiento"), 1):
                    ws2.cell(row=1, column=col_idx, value=col).font = Font(bold=True)

                for row_idx, (index, row) in enumerate(terceros_df.iterrows(), 2):
                    ws2.cell(row=row_idx, column=1, value=index)
                    for col_idx, value in enumerate(row, 2):
                        ws2.cell(row=row_idx, column=col_idx, value=value)

            bio = BytesIO()
            wb.save(bio)
            bio.seek(0)
            return bio


        bio_mensual = build_excel_mensual(pivot_propios, pivot_terceros)

        st.download_button(
            "📥 Descargar honorarios mensuales (Excel)",
            data=bio_mensual,
            file_name="honorarios_mensuales.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        st.subheader("🏠 Ranking por apartamento")
        st.dataframe(ranking, use_container_width=True)


