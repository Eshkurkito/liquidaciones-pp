import streamlit as st
import pandas as pd
import numpy as np
from datetime import date, timedelta
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from pathlib import Path
import re

st.set_page_config(page_title="LIQUIDACIONES (Casos 1–5) + Conciliación bancaria avanzada [v9 estricto]", page_icon="🏦", layout="wide")

# ========= Utilidades de formato =========
MONEY_COLS_CANON = {
    "Ingreso alojamiento","Ingreso limpieza","Total ingresos","Comisión portal",
    "Honorarios Florit","Gasto limpieza","Amenities","Total Gastos","Pago al propietario",
    "Pago recibido","IVA del alquiler"
}
NIGHTS_COL = "Noches ocupadas"

def ensure_unique_columns(df: pd.DataFrame) -> pd.DataFrame:
    counts, new_cols = {}, []
    for c in df.columns:
        name = str(c)
        n = counts.get(name, 0)
        new_cols.append(name if n == 0 else f"{name}.{n}")
        counts[name] = n + 1
    out = df.copy()
    out.columns = new_cols
    return out

def _first_existing(df, candidates):
    norm_map = {str(c).strip().lower(): c for c in df.columns}
    for cand in candidates:
        k = str(cand).strip().lower()
        if k in norm_map:
            return norm_map[k]
    return None

def ensure_required(df, required, ctx=""):
    missing = [c for c in required if c not in df.columns]
    if missing:
        st.error(f"Faltan columnas requeridas: {missing} en {ctx}. Ajusta el archivo o usa el modo por letras.")
        st.stop()

def base_name(colname: str) -> str:
    return re.sub(r"\.\d+$", "", str(colname)).strip()

def is_money_col(colname: str) -> bool:
    return base_name(colname) in MONEY_COLS_CANON

def is_nights_col(colname: str) -> bool:
    return base_name(colname).lower() == NIGHTS_COL.lower()

def fmt_number_for_ui(colname: str, x):
    if is_nights_col(colname):
        try:
            return f"{int(round(float(x)))}"
        except Exception:
            return x
    if is_money_col(colname):
        try:
            s = f"{float(x):,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
            return f"{s} €"
        except Exception:
            return x
    try:
        return f"{float(x):,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
    except Exception:
        return x

def find_col(df: pd.DataFrame, base: str):
    for c in df.columns:
        if base_name(c).lower() == base.strip().lower():
            return c
    return None

def show_table_es_grouped(df: pd.DataFrame, title: str, group_col: str = "Alojamiento"):
    st.subheader(title)

    if group_col not in df.columns:
        view = df.copy()
        total = {c: (view[c].sum() if pd.api.types.is_numeric_dtype(view[c]) else "") for c in view.columns}
        view = pd.concat([view, pd.DataFrame([total], index=["TOTAL"])], axis=0)

        view_fmt = view.copy()
        for c in view_fmt.columns:
            if pd.api.types.is_numeric_dtype(view[c]) or is_money_col(c) or is_nights_col(c):
                view_fmt[c] = view_fmt[c].apply(lambda v: fmt_number_for_ui(c, v))

        def highlight_total(row):
            return ["font-weight: bold;" if row.name == "TOTAL" else "" for _ in row]

        st.dataframe(view_fmt.style.apply(highlight_total, axis=1), use_container_width=True)
        return

    for aloj, subdf in df.groupby(group_col):
        st.markdown(f"**{aloj}**")
        block = subdf.copy()
        total = {c: (block[c].sum() if pd.api.types.is_numeric_dtype(block[c]) else "") for c in block.columns}
        block = pd.concat([block, pd.DataFrame([total], index=["TOTAL"])], axis=0)

        block_fmt = block.copy()
        for c in block_fmt.columns:
            if pd.api.types.is_numeric_dtype(block[c]) or is_money_col(c) or is_nights_col(c):
                block_fmt[c] = block_fmt[c].apply(lambda v: fmt_number_for_ui(c, v))

        def highlight_total(row):
            return ["font-weight: bold;" if row.name == "TOTAL" else "" for _ in row]

        st.dataframe(block_fmt.style.apply(highlight_total, axis=1), use_container_width=True)
        st.divider()

# ========= Normalización =========
LETTER_MAP_DEFAULT = {
    "W": "Alojamiento",
    "D": "Fecha entrada",
    "F": "Fecha salida",
    "H": "Noches ocupadas",
    "I": "Ingreso alojamiento",
    "J/L": "Ingreso limpieza",    # mapeo fuerte: tarifa limpieza en L
    "O": "Total ingresos",
    "AP": "Portal",
    "AR": "Comisión portal",
    "AL": "IVA del alquiler",
}

def letters_to_idx(letter):
    s = letter.upper()
    n = 0
    for ch in s:
        if not ('A' <= ch <= 'Z'): return None
        n = n*26 + (ord(ch)-ord('A')+1)
    return n-1

def normalize_columns_by_letters(df, letter_map=LETTER_MAP_DEFAULT):
    out = df.copy()
    cols = list(out.columns)
    rename = {}
    for L, std in letter_map.items():
        i = letters_to_idx(L)
        if i is not None and i < len(cols):
            rename[cols[i]] = std
    out.rename(columns=rename, inplace=True)
    return normalize_columns(out)

def normalize_columns(df):
    out = df.copy()
    col_aloj = _first_existing(out, ["Nombre alojamiento","Alojamiento","Nombre del alojamiento","Nombre Alojamiento"])
    col_fent = _first_existing(out, ["Fecha entrada","Fecha de entrada"])
    col_fsal = _first_existing(out, ["Fecha salida","Fecha de salida"])
    col_noch = _first_existing(out, ["Noches","noches","Noches ocupadas"])
    col_alq  = _first_existing(out, ["Alquiler con tasas","Ingreso alojamiento","Importe alojamiento"])
    col_ext  = _first_existing(out, [
        "Ingreso limpieza","Tarifa limpieza","Limpieza","Importe limpieza",
        "Extras con tasas","Gastos de limpieza","Gasto limpieza"
    ])
    col_tot  = _first_existing(out, ["Total reserva con tasas","Total ingresos","Total"])
    col_port = _first_existing(out, ["Web origen","Portal","Canal","Fuente"])
    col_comi = _first_existing(out, ["Comisión Portal/Intermediario: Comisión calculada","Comisión portal","Comisión"])
    col_ivaal= _first_existing(out, ["IVA del alojamiento","IVA alojamiento","IVA del alquiler"])

    rename = {}
    if col_aloj: rename[col_aloj] = "Alojamiento"
    if col_fent: rename[col_fent] = "Fecha entrada"
    if col_fsal: rename[col_fsal] = "Fecha salida"
    if col_noch: rename[col_noch] = "Noches ocupadas"
    if col_alq:  rename[col_alq]  = "Ingreso alojamiento"
    if col_ext:  rename[col_ext]  = "Ingreso limpieza"
    if col_tot:  rename[col_tot]  = "Total ingresos"
    if col_port: rename[col_port] = "Portal"
    if col_comi: rename[col_comi] = "Comisión portal"
    if col_ivaal:rename[col_ivaal]= "IVA del alquiler"

    out.rename(columns=rename, inplace=True)

    # Tipado
    for c in ["Ingreso alojamiento","Ingreso limpieza","Total ingresos","Comisión portal","IVA del alquiler","Noches ocupadas"]:
        if c in out.columns:
            out[c] = pd.to_numeric(out[c], errors="coerce").fillna(0.0)

    for c in ["Fecha entrada","Fecha salida"]:
        if c in out.columns:
            out[c] = pd.to_datetime(out[c], errors="coerce", dayfirst=True)

    if "Alojamiento" in out.columns:
        out["Alojamiento"] = out["Alojamiento"].astype(str).str.strip().str.upper()
    if "Noches ocupadas" in out.columns:
        out["Noches ocupadas"] = pd.to_numeric(out["Noches ocupadas"], errors="coerce").fillna(0).round(0).astype(int)

    return out

# ========= Reglas de casos (leer desde CSV) =========
# Se espera un archivo 'reglas_apartamentos.csv' junto al script o en la carpeta del proyecto.
_rules_path = Path(__file__).with_name("reglas_apartamentos.csv")
if not _rules_path.exists():
    _rules_path = Path(r"c:\Users\Usuario\Desktop\liquidaciones-pp\reglas_apartamentos.csv")
try:
    _rules_df = pd.read_csv(_rules_path)
except Exception:
    _rules_df = pd.DataFrame(columns=[
        "property","honorarios_pct","honorarios_apply_vat","honorarios_vat_pct",
        "amenities_amount","cleaning_fee","compute_iva_alquiler","commission_vat_pct",
        "treat_empty_portal_as_booking","skip_booking_vat","split_commission",
        "hon_base_exclude_commission","notes"
    ])

_rules_df.fillna("", inplace=True)

def _to_float(x, default=0.0):
    try:
        return float(x)
    except Exception:
        return default

props_rules = {}
for _, row in _rules_df.iterrows():
    prop = str(row.get("property", "")).strip().upper()
    if not prop:
        continue
    props_rules[prop] = {
        "honorarios_pct": _to_float(row.get("honorarios_pct", 0.20)),
        "honorarios_apply_vat": int(row.get("honorarios_apply_vat", 0)) if str(row.get("honorarios_apply_vat","")).strip()!="" else 0,
        "honorarios_vat_pct": _to_float(row.get("honorarios_vat_pct", 21.0)),
        "amenities_amount": _to_float(row.get("amenities_amount", 0.0)),
        "cleaning_fee": _to_float(row.get("cleaning_fee", 0.0)),
        "compute_iva_alquiler": int(row.get("compute_iva_alquiler", 0)) if str(row.get("compute_iva_alquiler","")).strip()!="" else 0,
        "commission_vat_pct": _to_float(row.get("commission_vat_pct", 0.0)),
        "treat_empty_portal_as_booking": str(row.get("treat_empty_portal_as_booking","")).strip() in ("1","True","true", "YES", "Yes"),
        "skip_booking_vat": str(row.get("skip_booking_vat","")).strip() in ("1","True","true", "YES", "Yes"),
        "split_commission": str(row.get("split_commission","")).strip() in ("1","True","true", "YES", "Yes"),
        "hon_base_exclude_commission": str(row.get("hon_base_exclude_commission","")).strip() in ("1","True","true", "YES", "Yes"),
        "notes": str(row.get("notes","")).strip()
    }

# Generar sets por "Caso" leyendo la columna notes (espera "Caso N" en notes)
case1_props = {p for p, v in props_rules.items() if "CASO 1" in v["notes"].upper()}
case2_props = {p for p, v in props_rules.items() if "CASO 2" in v["notes"].upper()}
case3_props = {p for p, v in props_rules.items() if "CASO 3" in v["notes"].upper()}
case4_props = {p for p, v in props_rules.items() if "CASO 4" in v["notes"].upper()}
case5_props = {p for p, v in props_rules.items() if "CASO 5" in v["notes"].upper()}

# APOLO_ONLY: propiedades marcadas con "APOLO" en nombre o en notes
APOLO_ONLY = {p for p in props_rules.keys() if "APOLO" in p or "APOLO" in props_rules[p]["notes"].upper()}

def props_for_case(case):
    if case == 1: return case1_props
    if case == 2: return case2_props
    if case == 3: return case3_props
    if case == 4: return case4_props
    if case == 5: return case5_props
    return set()

# ========= Reglas transversales =========
def apply_commission_vat_by_scope(df: pd.DataFrame, vat_pct: float, treat_empty_as_booking: bool, skip_booking_vat: bool,
                                  scope_mask: pd.Series | None = None) -> tuple[pd.DataFrame, int]:
    out = df.copy()
    portal_col = "Portal"; commission_col = "Comisión portal"
    if portal_col not in out.columns or commission_col not in out.columns:
        return out, 0

    ser_portal = out[portal_col]
    if isinstance(ser_portal, pd.DataFrame):
        ser_portal = ser_portal.iloc[:, 0]
    ser_portal = ser_portal.astype("string").fillna("")

    out[commission_col] = pd.to_numeric(out[commission_col], errors="coerce").fillna(0.0)
    mask_booking = ser_portal.str.lower().str.contains("booking", na=False)
    mask_empty   = ser_portal.str.strip().eq("")
    warn_count = int(((mask_empty) & (out[commission_col] > 0)).sum())

    if skip_booking_vat or vat_pct == 0:
        return out, warn_count

    mult = 1 + (float(vat_pct) / 100.0)
    mask_scope = scope_mask if scope_mask is not None else pd.Series(True, index=out.index)
    out.loc[(mask_booking & mask_scope), commission_col] *= mult
    if treat_empty_as_booking:
        out.loc[(mask_empty & mask_scope), commission_col] *= mult

    return out, warn_count

# ========= Procesadores =========
def process_case1(df, treat_empty_as_booking=False, skip_booking_vat=False, vat_pct=21.0):
    df = normalize_columns(df)
    ensure_required(df, ["Alojamiento","Ingreso alojamiento","Ingreso limpieza","Total ingresos","Comisión portal","Portal"], "Caso 1")
    scope = pd.Series(True, index=df.index)
    df, warn_count = apply_commission_vat_by_scope(df, vat_pct, treat_empty_as_booking, skip_booking_vat, scope)

    def honorarios(r):
        key = str(r.get("Alojamiento","")).strip().upper()
        pct = props_rules.get(key, {}).get("honorarios_pct", 0.20)
        # si honorarios_apply_vat está activado en reglas, aplicar 1 + vat%
        apply_v = props_rules.get(key, {}).get("honorarios_apply_vat", 1)
        vat_local = props_rules.get(key, {}).get("honorarios_vat_pct", 21.0)
        mult = 1.0 + (vat_local/100.0) if apply_v else 1.0
        return float(r.get("Ingreso alojamiento",0.0)) * pct * mult

    def amenities(r):
        key = str(r.get("Alojamiento","")).strip().upper()
        return float(props_rules.get(key, {}).get("amenities_amount", 0.0))

    out = df.copy()
    out["Honorarios Florit"] = out.apply(honorarios, axis=1).round(2)
    out["Gasto limpieza"]   = pd.to_numeric(out.get("Ingreso limpieza", 0.0), errors="coerce").fillna(0.0).round(2)
    out["Amenities"]        = out.apply(amenities, axis=1).round(2)
    out["Total Gastos"]     = (out[["Comisión portal","Honorarios Florit","Gasto limpieza","Amenities"]].sum(axis=1)).round(2)
    out["Pago al propietario"] = (out["Total ingresos"] - out["Total Gastos"]).round(2)
    out["Pago recibido"]    = (out["Total ingresos"] - out["Comisión portal"]).round(2)

    cols = ["Alojamiento","Fecha entrada","Fecha salida","Noches ocupadas","Ingreso alojamiento","Ingreso limpieza",
            "Total ingresos","Portal","Comisión portal","Honorarios Florit","Gasto limpieza","Amenities",
            "Total Gastos","Pago al propietario","Pago recibido"]
    cols = [c for c in cols if c in out.columns]
    return out[cols], warn_count

def process_case2(df, treat_empty_as_booking=False, skip_booking_vat=False, vat_pct=21.0, only_apolo=True):
    df = normalize_columns(df)
    ensure_required(df, ["Alojamiento","Ingreso alojamiento","Total ingresos","Comisión portal","Portal"], "Caso 2")
    mask_apolo = df["Alojamiento"].astype(str).str.upper().isin(APOLO_ONLY) if only_apolo else pd.Series(True, index=df.index)
    df, warn_count = apply_commission_vat_by_scope(df, vat_pct, treat_empty_as_booking, skip_booking_vat, scope_mask=mask_apolo)

    def honorarios(r):
        key = str(r.get("Alojamiento","")).strip().upper()
        pct = props_rules.get(key, {}).get("honorarios_pct", 0.20)
        ingreso = float(r.get("Ingreso alojamiento",0.0))
        # Si la regla indica compute_iva_alquiler, calculamos IVA del alquiler como antes, sino usamos default 1.10
        if props_rules.get(key, {}).get("compute_iva_alquiler", 0):
            iva = ingreso - (ingreso / 1.10)
            base = ingreso - iva
        else:
            iva = ingreso - (ingreso / 1.10)
            base = ingreso - iva
        apply_v = props_rules.get(key, {}).get("honorarios_apply_vat", 1)
        vat_local = props_rules.get(key, {}).get("honorarios_vat_pct", 21.0)
        mult = 1.0 + (vat_local/100.0) if apply_v else 1.0
        return base * pct * mult

    def amenities(r):
        key = str(r.get("Alojamiento","")).strip().upper()
        return float(props_rules.get(key, {}).get("amenities_amount", 0.0))

    out = df.copy()
    out["IVA del alquiler"] = pd.to_numeric(out["Ingreso alojamiento"], errors="coerce").fillna(0.0) - (pd.to_numeric(out["Ingreso alojamiento"], errors="coerce").fillna(0.0) / 1.10)
    out["Honorarios Florit"] = out.apply(honorarios, axis=1).round(2)
    out["Gasto limpieza"]   = pd.to_numeric(out.get("Ingreso limpieza", 0.0), errors="coerce").fillna(0.0).round(2)
    out["Amenities"]        = out.apply(amenities, axis=1).round(2)
    out["Total Gastos"]     = (out[["Comisión portal","Honorarios Florit","Gasto limpieza","Amenities"]].sum(axis=1)).round(2)
    out["Pago al propietario"] = (out["Total ingresos"] - out["Total Gastos"]).round(2)
    out["Pago recibido"]    = (out["Total ingresos"] - out["Comisión portal"]).round(2)

    cols = ["Alojamiento","Fecha entrada","Fecha salida","Noches ocupadas","Ingreso alojamiento","IVA del alquiler",
            "Ingreso limpieza","Total ingresos","Portal","Comisión portal","Honorarios Florit","Gasto limpieza",
            "Amenities","Total Gastos","Pago al propietario","Pago recibido"]
    cols = [c for c in cols if c in out.columns]
    return out[cols], warn_count

def process_case3(df, treat_empty_as_booking=False, skip_booking_vat=False, vat_pct=21.0):
    df = normalize_columns(df)
    ensure_required(df, ["Alojamiento","Ingreso alojamiento","Total ingresos","Comisión portal","Portal"], "Caso 3")
    scope = pd.Series(True, index=df.index)

    # En Caso 3 NO alteramos la comisión original; solo queremos el desglose.
    # Llamamos con skip_booking_vat=True para obtener warn_count pero sin tocar importes.
    df, warn_count = apply_commission_vat_by_scope(df, vat_pct, treat_empty_as_booking, True, scope)

    out = df.copy()

    # Desglose de comisión del portal
    comi_sin = pd.to_numeric(out.get("Comisión portal", 0.0), errors="coerce").fillna(0.0).round(2)
    out["Comisión portal (sin IVA)"] = comi_sin
    out["IVA comisión portal"] = (comi_sin * (float(vat_pct) / 100.0)).round(2)
    # “Comisión portal” pasa a ser CON IVA para que totales/pago recibido cuadren
    out["Comisión portal"] = (out["Comisión portal (sin IVA)"] + out["IVA comisión portal"]).round(2)

    def honorarios(r):
        key = str(r.get("Alojamiento","")).strip().upper()
        pct = props_rules.get(key, {}).get("honorarios_pct", 0.20)
        # Nueva fórmula: (alojamiento - comisión SIN IVA) * pct * posible IVA de honorarios
        base = float(r.get("Ingreso alojamiento",0.0)) - float(r.get("Comisión portal (sin IVA)",0.0))
        apply_v = props_rules.get(key, {}).get("honorarios_apply_vat", 1)
        vat_local = props_rules.get(key, {}).get("honorarios_vat_pct", 21.0)
        mult = 1.0 + (vat_local/100.0) if apply_v else 1.0
        return base * pct * mult

    def gasto_limpieza(r):
        key = str(r.get("Alojamiento","")).strip().upper()
        return float(props_rules.get(key, {}).get("cleaning_fee", 0.0))

    def amenities(r):
        key = str(r.get("Alojamiento","")).strip().upper()
        return float(props_rules.get(key, {}).get("amenities_amount", 0.0))

    out["Honorarios Florit"] = out.apply(honorarios, axis=1).round(2)
    out["Gasto limpieza"]   = out.apply(gasto_limpieza, axis=1).round(2)
    out["Amenities"]        = out.apply(amenities, axis=1).round(2)

    # Total Gastos con comisión CON IVA
    out["Total Gastos"] = (out[["Comisión portal","Honorarios Florit","Gasto limpieza","Amenities"]].sum(axis=1)).round(2)

    out["Pago al propietario"] = (out["Total ingresos"] - out["Total Gastos"]).round(2)
    out["Pago recibido"]       = (out["Total ingresos"] - out["Comisión portal"]).round(2)

    cols = ["Alojamiento","Fecha entrada","Fecha salida","Noches ocupadas",
            "Ingreso alojamiento","Ingreso limpieza","Total ingresos","Portal",
            "Comisión portal (sin IVA)","IVA comisión portal","Comisión portal",
            "Honorarios Florit","Gasto limpieza","Amenities",
            "Total Gastos","Pago al propietario","Pago recibido"]
    cols = [c for c in cols if c in out.columns]
    return out[cols], warn_count

def process_case4(df, treat_empty_as_booking=False, skip_booking_vat=False, vat_pct=0.0):
    df = normalize_columns(df)
    ensure_required(df, ["Alojamiento","Ingreso alojamiento","Comisión portal"], "Caso 4")
    df["Portal"] = df.get("Portal", pd.Series([""]*len(df)))
    scope = pd.Series(True, index=df.index)
    df, warn_count = apply_commission_vat_by_scope(df, vat_pct, treat_empty_as_booking, skip_booking_vat, scope)

    out = df.copy()
    ingreso = pd.to_numeric(out.get("Ingreso alojamiento", 0.0), errors="coerce").fillna(0.0)
    out["IVA del alquiler"] = ingreso - (ingreso / 1.10)

    def honorarios(r):
        base = float(r.get("Ingreso alojamiento",0.0)) - float(r.get("IVA del alquiler",0.0)) - float(r.get("Comisión portal",0.0))
        return base * 0.20

    out["Honorarios Florit"] = out.apply(honorarios, axis=1).round(2)

    out["Pago al propietario"] = (
        pd.to_numeric(out.get("Ingreso alojamiento",0.0), errors="coerce").fillna(0.0)
        - pd.to_numeric(out.get("IVA del alquiler",0.0), errors="coerce").fillna(0.0)
        - pd.to_numeric(out.get("Comisión portal",0.0), errors="coerce").fillna(0.0)
        - pd.to_numeric(out.get("Honorarios Florit",0.0), errors="coerce").fillna(0.0)
    ).round(2)

    out["Pago recibido"] = (
        pd.to_numeric(out.get("Ingreso alojamiento",0.0), errors="coerce").fillna(0.0)
        + pd.to_numeric(out.get("Ingreso limpieza",0.0), errors="coerce").fillna(0.0)
        - pd.to_numeric(out.get("Comisión portal",0.0), errors="coerce").fillna(0.0)
    ).round(2)

    cols = [
        "Alojamiento","Fecha entrada","Fecha salida","Noches ocupadas",
        "Ingreso alojamiento","IVA del alquiler","Ingreso limpieza","Total ingresos",
        "Portal","Comisión portal","Honorarios Florit","Pago al propietario","Pago recibido"
    ]
    cols = [c for c in cols if c in out.columns]
    return out[cols], warn_count

def process_case5(df, treat_empty_as_booking=False, skip_booking_vat=False, vat_pct=0.0):
    df = normalize_columns(df)
    ensure_required(df, ["Alojamiento","Ingreso alojamiento","Total ingresos","Comisión portal"], "Caso 5")
    df["Portal"] = df.get("Portal", pd.Series([""]*len(df)))
    scope = pd.Series(True, index=df.index)
    df, warn_count = apply_commission_vat_by_scope(df, vat_pct, treat_empty_as_booking, skip_booking_vat, scope)

    out = df.copy()
    ingreso = pd.to_numeric(out.get("Ingreso alojamiento", 0.0), errors="coerce").fillna(0.0)
    out["IVA del alquiler"] = ingreso - (ingreso / 1.10)

    def honorarios(r):
        key = str(r.get("Alojamiento","")).strip().upper()
        pct = props_rules.get(key, {}).get("honorarios_pct", 0.20)
        base = float(r.get("Ingreso alojamiento",0.0)) - float(r.get("IVA del alquiler",0.0)) - float(r.get("Comisión portal",0.0))
        apply_v = props_rules.get(key, {}).get("honorarios_apply_vat", 1)
        vat_local = props_rules.get(key, {}).get("honorarios_vat_pct", 21.0)
        mult = 1.0 + (vat_local/100.0) if apply_v else 1.0
        return base * pct * mult

    def amenities(r):
        key = str(r.get("Alojamiento","")).strip().upper()
        return float(props_rules.get(key, {}).get("amenities_amount", 0.0))

    out["Honorarios Florit"] = out.apply(honorarios, axis=1).round(2)
    out["Gasto limpieza"]   = pd.to_numeric(out.get("Ingreso limpieza", 0.0), errors="coerce").fillna(0.0).round(2)
    out["Amenities"]        = out.apply(amenities, axis=1).round(2)
    out["Total Gastos"]     = (out[["Comisión portal","Honorarios Florit","Gasto limpieza","Amenities"]].sum(axis=1)).round(2)
    out["Pago al propietario"] = (out["Total ingresos"] - out["Total Gastos"]).round(2)
    out["Pago recibido"]    = (out["Total ingresos"] - out["Comisión portal"]).round(2)

    cols = [
        "Alojamiento","Fecha entrada","Fecha salida","Noches ocupadas",
        "Ingreso alojamiento","IVA del alquiler","Ingreso limpieza","Total ingresos","Portal","Comisión portal",
        "Honorarios Florit","Gasto limpieza","Amenities","Total Gastos","Pago al propietario","Pago recibido"
    ]
    cols = [c for c in cols if c in out.columns]
    return out[cols], warn_count

processors = {1: process_case1, 2: process_case2, 3: process_case3, 4: process_case4, 5: process_case5}

# ========= Exportación Excel =========
BORDER_THIN = Border(left=Side(style="thin"), right=Side(style="thin"),
                     top=Side(style="thin"), bottom=Side(style="thin"))

def write_grouped_sheet(ws, df):
    cols = list(df.columns)

    def write_table(start_row, subdf):
        # Cabecera
        for j, col in enumerate(cols, start=1):
            cell = ws.cell(row=start_row, column=j, value=col)
            cell.font = Font(bold=True)
            cell.border = BORDER_THIN
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        # Filas
        for i, (_, row) in enumerate(subdf.iterrows(), start=1):
            for j, col in enumerate(cols, start=1):
                val = row[col]
                c = ws.cell(row=start_row+i, column=j, value=val)
                c.border = BORDER_THIN
                if isinstance(val, (int, float)) and not pd.isna(val):
                    if is_nights_col(col):
                        c.number_format = "0"
                    elif is_money_col(col):
                        c.number_format = '#.##0,00" €"'
                    else:
                        c.number_format = "#.##0,00"
                else:
                    c.alignment = Alignment(wrap_text=True)
        # Sumatorios en negrita
        sum_row = start_row + len(subdf) + 1
        ws.cell(row=sum_row, column=1, value="TOTAL").font = Font(bold=True)
        ws.cell(row=sum_row, column=1).border = BORDER_THIN
        for j, col in enumerate(cols, start=1):
            if j == 1:
                continue
            if pd.api.types.is_numeric_dtype(subdf[col]):
                top = start_row+1
                bottom = start_row+len(subdf)
                formula = f"=SUM({get_column_letter(j)}{top}:{get_column_letter(j)}{bottom})"
                c = ws.cell(row=sum_row, column=j, value=formula)
                c.font = Font(bold=True)
                c.border = BORDER_THIN
                if is_nights_col(col):
                    c.number_format = "0"
                elif is_money_col(col):
                    c.number_format = '#.##0,00" €"'
                else:
                    c.number_format = "#.##0,00"
            else:
                ws.cell(row=sum_row, column=j, value="").border = BORDER_THIN
        return sum_row + 2

    current_row = 1
    if "Alojamiento" in df.columns:
        for aloj, subdf in df.groupby("Alojamiento"):
            ws.cell(row=current_row, column=1, value=str(aloj)).font = Font(bold=True, size=12)
            current_row += 1
            current_row = write_table(current_row, subdf)
    else:
        current_row = write_table(current_row, df)

    # Auto-ancho
    for j, col in enumerate(cols, start=1):
        max_len = len(str(col))
        for r in range(1, ws.max_row+1):
            v = ws.cell(row=r, column=j).value
            if v is not None:
                max_len = max(max_len, len(str(v)))
        ws.column_dimensions[get_column_letter(j)].width = min(max_len+2, 45)

def build_excel_single(df_final, filename="Liquidacion.xlsx"):
    wb = Workbook(); ws = wb.active; ws.title = "Liquidación"
    write_grouped_sheet(ws, df_final)
    bio = BytesIO(); wb.save(bio); bio.seek(0)
    st.download_button("📥 Descargar Excel (Liquidación)", bio.getvalue(),
                       file_name=filename,
                       mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

def build_excel_multi(dfs_by_case: dict, filename: str):
    wb = Workbook(); first = True
    for case_label, df_final in dfs_by_case.items():
        if first:
            ws = wb.active; ws.title = case_label; first = False
        else:
            ws = wb.create_sheet(title=case_label)
        write_grouped_sheet(ws, df_final)
    bio = BytesIO(); wb.save(bio); bio.seek(0)
    st.download_button("📥 Descargar Excel (Todos los casos)", bio.getvalue(),
                       file_name=filename,
                       mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

# ========= UI: LIQUIDACIONES =========
st.title("📊 LIQUIDACIONES (Casos 1–5)")
st.caption("Primero genera las liquidaciones del período. Luego sube el extracto bancario y concilia.")

with st.sidebar:
    st.header("Parámetros de liquidación")
    c1, c2 = st.columns(2)
    with c1:
        start_date = st.date_input("Desde", value=date(date.today().year, date.today().month, 1))
    with c2:
        end_date   = st.date_input("Hasta",  value=date(date.today().year, date.today().month, 28))
    st.divider()
    case_choice = st.radio("Caso", ["Todos", 1,2,3,4,5], horizontal=False)
    st.checkbox("Lectura por letras (fallback)", value=False, key="by_letters")
    st.caption("Mapeo: W, D, F, H, I, J/L (L limpia), O, AP, AR, AL.")
    st.divider()

    # Ajustes adicionales ocultables con "ojito"
    with st.expander("👁️ Ajustes adicionales", expanded=False):
        st.subheader("IVA comisión por caso (Booking)")
        col_v1, col_v2 = st.columns(2)
        with col_v1:
            vat_case1 = st.number_input("Caso 1 (%)", min_value=0.0, max_value=30.0, value=21.0, step=0.5)
            vat_case3 = st.number_input("Caso 3 (%)", min_value=0.0, max_value=30.0, value=21.0, step=0.5)
            vat_case5 = st.number_input("Caso 5 (%)", min_value=0.0, max_value=30.0, value=0.0, step=0.5)
        with col_v2:
            vat_case2 = st.number_input("Caso 2 (%)", min_value=0.0, max_value=30.0, value=21.0, step=0.5)
            vat_case4 = st.number_input("Caso 4 (%)", min_value=0.0, max_value=30.0, value=0.0, step=0.5)
            only_apolo_c2 = st.checkbox("Caso 2: aplicar solo a APOLO 029/197", value=True)
        treat_empty_as_booking = st.checkbox("Tratar portal vacío como Booking (aplicar IVA comisión)", value=False)
        skip_booking_vat = st.checkbox("No añadir IVA a comisión de Booking (ya viene con IVA)", value=False)

    generate = st.button("Generar liquidación")

# Nuevo: opción para indicar que la cabecera del Excel está en la fila 2
header_second_row = st.checkbox("La cabecera está en la segunda fila (leer desde la fila 2)", value=False)

file = st.file_uploader("Sube el archivo de reservas (.xlsx)", type=["xlsx"], key="reservas_upl")

# ========= Generación Liquidaciones =========
def normalize_liq_for_period(df_norm, start_date, end_date):
    if "Fecha entrada" in df_norm.columns:
        mask = (df_norm["Fecha entrada"] >= pd.to_datetime(start_date)) & (df_norm["Fecha entrada"] <= pd.to_datetime(end_date))
        df_norm = df_norm[mask]
    return df_norm

if generate:
    if not file:
        st.error("Sube primero el archivo de reservas (.xlsx).")
        st.stop()

    header_row = 1 if header_second_row else 0
    df_in = pd.read_excel(file, header=header_row)
    df_in = ensure_unique_columns(df_in)
    df_norm = normalize_columns_by_letters(df_in) if st.session_state.by_letters else normalize_columns(df_in)
    df_norm = ensure_unique_columns(df_norm)
    df_norm = normalize_liq_for_period(df_norm, start_date, end_date)

    if "Ingreso limpieza" in df_norm.columns:
        limp = pd.to_numeric(df_norm["Ingreso limpieza"], errors="coerce").fillna(0)
        if (limp > 300).any():
            st.warning("Detectadas tarifas de limpieza > 300 €. Verifica que la columna L esté mapeada como 'Ingreso limpieza' o activa el modo por letras.")

    def run_case(case_no):
        df_case = df_norm.copy()
        props = props_for_case(case_no)
        if props and "Alojamiento" in df_case.columns:
            df_case = df_case[df_case["Alojamiento"].isin(props)]
        vat_map = {1: vat_case1, 2: vat_case2, 3: vat_case3, 4: vat_case4, 5: vat_case5}
        if case_no == 2:
            out, warn = processors[case_no](df_case, treat_empty_as_booking=treat_empty_as_booking, skip_booking_vat=skip_booking_vat, vat_pct=vat_map[case_no], only_apolo=only_apolo_c2)
        elif case_no in (1,3,4,5):
            out, warn = processors[case_no](df_case, treat_empty_as_booking=treat_empty_as_booking, skip_booking_vat=skip_booking_vat, vat_pct=vat_map[case_no])
        else:
            out, warn = processors[case_no](df_case)
        if NIGHTS_COL in out.columns:
            out[NIGHTS_COL] = pd.to_numeric(out[NIGHTS_COL], errors="coerce").fillna(0).round(0).astype(int)
        for c in out.columns:
            if c != NIGHTS_COL and pd.api.types.is_numeric_dtype(out[c]):
                out[c] = pd.to_numeric(out[c], errors="coerce").fillna(0.0).round(2)
        return out, warn

    if case_choice == "Todos":
        dfs = {}; total_warns = 0
        for c in [1,2,3,4,5]:
            df_out, warn = run_case(c)
            total_warns += warn
            df_out = df_out.sort_values(by=[col for col in ["Alojamiento","Fecha entrada"] if col in df_out.columns])
            dfs[f"Caso {c}"] = df_out
        st.success(f"Liquidación generada (Todos) • {start_date.strftime('%d/%m/%Y')}–{end_date.strftime('%d/%m/%Y')}")

        for label, df_show in dfs.items():
            show_table_es_grouped(df_show, f"{label} — Tabla de liquidaciones")

        file_name = f"Liquidaciones_TODOS_{start_date.strftime('%Y%m%d')}_{end_date.strftime('%Y%m%d')}.xlsx"
        build_excel_multi(dfs, filename=file_name)

        st.session_state["df_liq_all"] = pd.concat(dfs.values(), ignore_index=True, sort=False)
        st.session_state["df_liq_label"] = "Todos"
        if total_warns > 0 and not treat_empty_as_booking:
            st.warning(f"Hay {total_warns} reservas con comisión > 0 pero portal vacío en alguno de los casos. Si deben ser Booking, marca la opción correspondiente y vuelve a generar.")
    else:
        case_no = int(case_choice)
        df_out, warn = run_case(case_no)
        df_out = df_out.sort_values(by=[col for col in ["Alojamiento","Fecha entrada"] if col in df_out.columns])

        st.success(f"Liquidación generada (Caso {case_no}) • {start_date.strftime('%d/%m/%Y')}–{end_date.strftime('%d/%m/%Y')}")
        show_table_es_grouped(df_out, "Tabla de liquidaciones")

        aloj_col = find_col(df_out, "Alojamiento")
        pago_col = find_col(df_out, "Pago al propietario")
        if aloj_col is not None and pago_col is not None:
            pagos = (df_out[[aloj_col, pago_col]].groupby(aloj_col, as_index=False)[pago_col]
                     .sum().round(2).sort_values(aloj_col))
            pagos.rename(columns={aloj_col: "Alojamiento", pago_col: "Pago al propietario"}, inplace=True)
            pagos_fmt = pagos.copy()
            for c in pagos_fmt.columns:
                if pd.api.types.is_numeric_dtype(pagos_fmt[c]) or is_money_col(c):
                    pagos_fmt[c] = pagos_fmt[c].apply(lambda v: fmt_number_for_ui(c, v))
            st.subheader("💸 Pagos por alojamiento (suma)")
            st.dataframe(pagos_fmt, use_container_width=True)

        file_case_name = f"Liquidacion_CASO{case_no}_{start_date.strftime('%Y%m%d')}_{end_date.strftime('%Y%m%d')}.xlsx"
        build_excel_single(df_out, filename=file_case_name)

        st.session_state["df_liq_all"] = df_out.copy()
        st.session_state["df_liq_label"] = f"Caso {case_no}"
        if warn > 0 and not treat_empty_as_booking:
            st.warning("Hay reservas con comisión > 0 pero portal vacío. Si deben ser Booking, marca ‘Tratar portal vacío como Booking’.")